#!/usr/bin/env python3
"""
Smart Excel Router - Reference Implementation

Complete implementation for the excel-parser skill demonstrating both
Pandas Standard Mode and HTML Semantic Mode processing paths.

This file serves as a reference for understanding the implementation details.
The actual processing is typically done inline by the agent using the patterns shown here.

Usage:
    python smart_excel_router.py <file_path> [--sheet <name>] [--force-mode pandas|html]

Dependencies:
    - openpyxl
    - pandas

Author: Claude Code - excel-parser Skill
Version: 1.0
"""

import openpyxl
from openpyxl.utils import range_boundaries
import pandas as pd
import json
import sys
import argparse
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple


class SmartExcelRouter:
    """
    Intelligent Excel file processor that routes to optimal strategy
    based on file complexity analysis.
    """

    # Complexity thresholds
    DEEP_MERGE_THRESHOLD = 2      # Max deep merges before marking complex
    EMPTY_ROW_THRESHOLD = 2       # Max empty row interruptions
    LARGE_FILE_THRESHOLD = 1000   # Rows above this force Pandas mode
    HEADER_REGION_BOUNDARY = 5    # Rows 1-5 considered header region
    SAMPLE_SIZE = 20              # Rows to sample for header detection

    def __init__(self, file_path: str):
        """
        Initialize router with Excel file.

        Args:
            file_path: Path to Excel file (.xlsx, .xlsm)

        Raises:
            FileNotFoundError: If file does not exist
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # Load workbook for metadata access
        self.wb = openpyxl.load_workbook(
            str(self.file_path),
            read_only=False,  # Required for merged_cells access
            data_only=True    # Read calculated values, not formulas
        )

    def analyze_complexity(self, sheet_name: str) -> Dict[str, Any]:
        """
        Analyze sheet complexity without loading full data.

        This is the "scout" function that examines metadata to determine
        the optimal processing strategy.

        Args:
            sheet_name: Name of sheet to analyze

        Returns:
            dict: Analysis results including:
                - is_complex: bool
                - recommended_strategy: "pandas" | "html"
                - reasons: list of strings explaining decision
                - stats: dict of metrics

        Raises:
            ValueError: If sheet_name not found
        """
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        sheet = self.wb[sheet_name]

        # Gather basic metrics
        max_row = sheet.max_row
        max_col = sheet.max_column
        merged_ranges = sheet.merged_cells.ranges

        # Analyze merged cell distribution
        deep_merges = 0
        shallow_merges = 0

        for merge in merged_ranges:
            min_col, min_row, max_col_merge, max_row_merge = range_boundaries(str(merge))
            if min_row > self.HEADER_REGION_BOUNDARY:
                deep_merges += 1
            else:
                shallow_merges += 1

        # Check for empty row interruptions (only on smaller tables)
        empty_interruptions = 0
        if max_row < 200:
            empty_interruptions = self._count_empty_interruptions(sheet, max_row)

        # Apply scoring rules
        is_complex = False
        reasons = []

        # Rule 1: Deep merged cells indicate complex structure
        if deep_merges > self.DEEP_MERGE_THRESHOLD:
            is_complex = True
            reasons.append(f"Detected {deep_merges} merged cells in data region")

        # Rule 2: Multiple empty rows suggest multi-table layout
        if max_row < 300 and empty_interruptions > self.EMPTY_ROW_THRESHOLD:
            is_complex = True
            reasons.append(f"Detected {empty_interruptions} empty row interruptions (multi-table layout)")

        # Rule 3: Large files must use Pandas (overrides complexity)
        if max_row > self.LARGE_FILE_THRESHOLD:
            is_complex = False
            reasons = [f"Row count ({max_row}) exceeds {self.LARGE_FILE_THRESHOLD}, forcing Pandas mode"]

        # Rule 4: Default for standard tables
        if not is_complex and not reasons:
            reasons.append("Standard table structure, no complex patterns detected")

        return {
            "is_complex": is_complex,
            "recommended_strategy": "html" if is_complex else "pandas",
            "reasons": reasons,
            "stats": {
                "total_rows": max_row,
                "total_columns": max_col,
                "deep_merges": deep_merges,
                "shallow_merges": shallow_merges,
                "empty_interruptions": empty_interruptions
            }
        }

    def _count_empty_interruptions(self, sheet, max_row: int) -> int:
        """
        Count transitions from data rows to empty rows to data rows.

        Args:
            sheet: openpyxl worksheet
            max_row: Maximum row number to check

        Returns:
            int: Number of empty row interruptions
        """
        empty_interruptions = 0
        consecutive_empty = 0

        for row_idx in range(1, max_row + 1):
            row = list(sheet.iter_rows(
                min_row=row_idx,
                max_row=row_idx,
                values_only=True
            ))[0]

            is_empty = all(cell is None or str(cell).strip() == '' for cell in row)

            if is_empty:
                consecutive_empty += 1
            else:
                if consecutive_empty > 0 and row_idx > 1:
                    empty_interruptions += 1
                consecutive_empty = 0

        return empty_interruptions

    def process_pandas_mode(self, sheet_name: str, header_row: int = 0) -> pd.DataFrame:
        """
        Process sheet using Pandas Standard Mode (Path A).

        This mode is optimal for simple/large tables. It reads the full
        data at native Pandas speed after determining the header row.

        Args:
            sheet_name: Name of sheet to process
            header_row: Row index (0-based) containing column headers

        Returns:
            pd.DataFrame: Loaded data with proper column headers
        """
        print(f"[Pandas Mode] Processing sheet: {sheet_name}")

        # Read full data with specified header
        df = pd.read_excel(
            str(self.file_path),
            sheet_name=sheet_name,
            header=header_row
        )

        print(f"  Loaded DataFrame: {df.shape[0]} rows x {df.shape[1]} columns")
        return df

    def sample_for_header_detection(self, sheet_name: str) -> str:
        """
        Sample first N rows for header detection.

        This output can be analyzed to determine which row contains
        the actual column headers.

        Args:
            sheet_name: Name of sheet to sample

        Returns:
            str: CSV formatted sample of first SAMPLE_SIZE rows
        """
        df_sample = pd.read_excel(
            str(self.file_path),
            sheet_name=sheet_name,
            header=None,
            nrows=self.SAMPLE_SIZE
        )
        return df_sample.to_csv(index=False)

    def process_html_mode(self, sheet_name: str, max_rows: Optional[int] = None) -> str:
        """
        Process sheet using HTML Semantic Mode (Path B).

        This mode preserves table structure including merged cells
        using rowspan/colspan attributes. Best for complex tables
        with irregular layouts.

        Args:
            sheet_name: Name of sheet to process
            max_rows: Optional limit on rows to convert (for token management)

        Returns:
            str: HTML table representation with semantic structure
        """
        print(f"[HTML Mode] Processing sheet: {sheet_name}")

        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")

        sheet = self.wb[sheet_name]
        actual_max_row = min(sheet.max_row, max_rows) if max_rows else sheet.max_row

        # Build merge map for rowspan/colspan
        merge_map = self._build_merge_map(sheet)
        skip_cells = self._build_skip_set(sheet)

        # Generate HTML
        html_parts = ['<table border="1">']

        for row_idx in range(1, actual_max_row + 1):
            html_parts.append('  <tr>')

            for col_idx in range(1, sheet.max_column + 1):
                # Skip cells that are part of a merge (not origin)
                if (row_idx, col_idx) in skip_cells:
                    continue

                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value if cell.value is not None else ''

                # Add span attributes if this is a merge origin
                attrs = self._get_span_attrs(row_idx, col_idx, merge_map)

                html_parts.append(f'    <td{attrs}>{value}</td>')

            html_parts.append('  </tr>')

        html_parts.append('</table>')

        html_content = '\n'.join(html_parts)
        print(f"  Generated HTML: {len(html_content)} characters")

        return html_content

    def _build_merge_map(self, sheet) -> Dict[Tuple[int, int], Dict[str, int]]:
        """
        Build mapping of merge origins to their spans.

        Args:
            sheet: openpyxl worksheet

        Returns:
            dict: {(row, col): {"rowspan": n, "colspan": m}}
        """
        merge_map = {}

        for merge in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merge))
            merge_map[(min_row, min_col)] = {
                'rowspan': max_row - min_row + 1,
                'colspan': max_col - min_col + 1
            }

        return merge_map

    def _build_skip_set(self, sheet) -> set:
        """
        Build set of cells to skip (non-origin cells in merged ranges).

        Args:
            sheet: openpyxl worksheet

        Returns:
            set: {(row, col), ...} cells to skip
        """
        skip_cells = set()

        for merge in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merge))

            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if (r, c) != (min_row, min_col):  # Skip non-origin cells
                        skip_cells.add((r, c))

        return skip_cells

    def _get_span_attrs(self, row: int, col: int,
                        merge_map: Dict[Tuple[int, int], Dict[str, int]]) -> str:
        """
        Get HTML span attributes for a cell.

        Args:
            row: Row index
            col: Column index
            merge_map: Merge origin mapping

        Returns:
            str: HTML attributes string (e.g., ' rowspan="2" colspan="3"')
        """
        if (row, col) not in merge_map:
            return ''

        span = merge_map[(row, col)]
        attrs = ''

        if span['rowspan'] > 1:
            attrs += f' rowspan="{span["rowspan"]}"'
        if span['colspan'] > 1:
            attrs += f' colspan="{span["colspan"]}"'

        return attrs

    def process_auto(self, sheet_name: str) -> Dict[str, Any]:
        """
        Automatically analyze and process sheet with optimal strategy.

        Args:
            sheet_name: Name of sheet to process

        Returns:
            dict: {
                "strategy": "pandas" | "html",
                "analysis": {...},
                "data": DataFrame or HTML string
            }
        """
        analysis = self.analyze_complexity(sheet_name)

        if analysis["recommended_strategy"] == "pandas":
            # For demonstration, assume header at row 0
            # In practice, sample and analyze first
            data = self.process_pandas_mode(sheet_name, header_row=0)
        else:
            data = self.process_html_mode(sheet_name)

        return {
            "strategy": analysis["recommended_strategy"],
            "analysis": analysis,
            "data": data
        }

    def close(self):
        """Close the workbook."""
        self.wb.close()


def main():
    """CLI entry point."""
    parser = argparse.ArgumentParser(
        description="Smart Excel Router - Intelligent file processing"
    )
    parser.add_argument("file_path", help="Path to Excel file")
    parser.add_argument("--sheet", help="Sheet name (default: first sheet)")
    parser.add_argument(
        "--force-mode",
        choices=["pandas", "html"],
        help="Force specific processing mode"
    )
    parser.add_argument(
        "--analyze-only",
        action="store_true",
        help="Only analyze complexity, don't process"
    )

    args = parser.parse_args()

    try:
        router = SmartExcelRouter(args.file_path)

        # Determine sheet to process
        sheet_name = args.sheet or router.wb.sheetnames[0]

        # Analyze complexity
        analysis = router.analyze_complexity(sheet_name)
        print(f"\n=== Complexity Analysis: {sheet_name} ===")
        print(json.dumps(analysis, indent=2))

        if args.analyze_only:
            router.close()
            return

        # Process based on mode
        mode = args.force_mode or analysis["recommended_strategy"]
        print(f"\n=== Processing with {mode.upper()} mode ===")

        if mode == "pandas":
            # Sample for header detection first
            sample = router.sample_for_header_detection(sheet_name)
            print("\nFirst 20 rows (for header detection):")
            print(sample[:500] + "..." if len(sample) > 500 else sample)

            # Process with default header (row 0)
            df = router.process_pandas_mode(sheet_name, header_row=0)
            print(f"\nDataFrame preview:\n{df.head()}")
        else:
            html = router.process_html_mode(sheet_name, max_rows=100)
            print(f"\nHTML output (truncated):\n{html[:1000]}...")

        router.close()

    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
